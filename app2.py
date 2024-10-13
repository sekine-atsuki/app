import threading
import time
import io
from flask import Flask, request, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import webview
import logging
import sys
import os

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # ファイルを取得
        file = request.files['file']
        if file:
            # pandasでExcelファイルを読み込む
            df = pd.read_excel(file, header=None)

            # ヘッダーとデータの分離
            header = df.iloc[:2]
            data = df.iloc[2:]

            # C列でソート
            data_sorted = data.sort_values(by=data.columns[2], ascending=True)

            # ヘッダーとソートされたデータを結合
            df_sorted = pd.concat([header, data_sorted], ignore_index=True)

            # pandasのデータフレームをExcelファイルに保存
            output_pandas = io.BytesIO()
            with pd.ExcelWriter(output_pandas, engine='openpyxl') as writer:
                df_sorted.to_excel(writer, index=False, header=False)
            output_pandas.seek(0)

            # openpyxlでExcelファイルを開く
            wb = load_workbook(output_pandas)
            ws = wb.active

            # 列幅の設定
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 7
            ws.column_dimensions['C'].width = 12.5
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15

            # フォントとセンタリングの設定
            font = Font(name='游ゴシック')
            alignment = Alignment(horizontal='center', vertical='center')

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment

            # 日付と曜日の設定
            today = datetime.today()
            japanese_weekdays = ['月', '火', '水', '木', '金', '土', '日']
            weekday_japanese = japanese_weekdays[today.weekday()]
            formatted_date = today.strftime(f'%m/%d（{weekday_japanese}）')

            # A1セルの設定
            ws['A1'] = f"{formatted_date}朝食リスト"
            ws.merge_cells('A1:D1')
            fill = PatternFill(start_color='CCF2CC', end_color='CCF2CC', fill_type='solid')
            ws['A1'].fill = fill
            ws['A1'].font = font
            ws['A1'].alignment = alignment

            # A列の最後に「計」を追加
            last_row = ws.max_row + 1
            ws[f'A{last_row}'] = "計"
            ws[f'A{last_row}'].font = font
            ws[f'A{last_row}'].alignment = alignment

            # B列の合計を計算
            b_values = [cell.value for cell in ws['B'] if isinstance(cell.value, (int, float))]
            ws[f'B{last_row}'] = sum(b_values)
            ws[f'B{last_row}'].font = font
            ws[f'B{last_row}'].alignment = alignment

            # 時間枠をF列に追加
            start_time = datetime.strptime("07:00", "%H:%M")
            end_time = datetime.strptime("10:00", "%H:%M")
            time_increment = timedelta(minutes=30)

            current_time = start_time
            row_num = 3
            while current_time <= end_time:
                time_range = f"{current_time.strftime('%H:%M')}~{(current_time + time_increment).strftime('%H:%M')}"
                ws[f'F{row_num}'] = time_range
                ws[f'F{row_num}'].font = font
                ws[f'F{row_num}'].alignment = alignment
                current_time += time_increment
                row_num += 1

            # F列の最後に「計」を追加
            ws[f'F{last_row}'] = "計"
            ws[f'F{last_row}'].font = font
            ws[f'F{last_row}'].alignment = alignment

            # 行の高さを設定
            for row in range(1, last_row + 1):
                ws.row_dimensions[row].height = 20

            # 時間枠ごとのB列の合計を計算し、G列に出力
            time_sums = {}
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=3):
                time_value = row[0].value
                if time_value not in time_sums:
                    time_sums[time_value] = 0
                b_value = ws[f'B{row[0].row}'].value
                if isinstance(b_value, (int, float)):
                    time_sums[time_value] += b_value

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=6, max_col=6):
                time_value = row[0].value
                if time_value in time_sums and time_sums[time_value] > 0:
                    ws[f'G{row[0].row}'] = time_sums[time_value]
                else:
                    ws[f'G{row[0].row}'] = 0
                ws[f'G{row[0].row}'].font = font
                ws[f'G{row[0].row}'].alignment = alignment

            # G列の合計を計算
            g_values = [cell.value for cell in ws['G'] if isinstance(cell.value, (int, float))]
            ws[f'G{last_row}'] = sum(g_values)
            ws[f'G{last_row}'].font = font
            ws[f'G{last_row}'].alignment = alignment

            # グレーの塗りつぶしを適用
            gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            ws[f'A{last_row}'].fill = gray_fill
            ws[f'B{last_row}'].fill = gray_fill
            ws[f'F{last_row}'].fill = gray_fill
            ws[f'G{last_row}'].fill = gray_fill

            # ボーダーを設定
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border

            # 斜線を適用
            diagonal_border = Border(
                diagonal=Side(style='thin'),
                diagonal_direction=1
            )
            ws[f'C{last_row}'].border = diagonal_border
            ws[f'D{last_row}'].border = diagonal_border

            # 処理済みのExcelファイルをバッファに保存
            output_final = io.BytesIO()
            wb.save(output_final)
            output_final.seek(0)

            # ファイルをダウンロードさせる
            return send_file(
                output_final,
                as_attachment=True,
                download_name='processed_file.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

    # アップロードフォームを表示
    return '''
    <!doctype html>
    <html>
    <head>
        <title>Excelファイルのアップロード</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 50px; }
            h1 { color: #333; }
            form { margin-top: 30px; }
            input[type="file"] { margin-bottom: 20px; }
            input[type="submit"] {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                cursor: pointer;
            }
            input[type="submit"]:hover {
                background-color: #45a049;
            }
        </style>
    </head>
    <body>
        <h1>Excelファイルをアップロードしてください</h1>
        <form action="/" method="post" enctype="multipart/form-data">
            <input type="file" name="file" required><br>
            <input type="submit" value="アップロード">
        </form>
    </body>
    </html>
    '''

def start_server():
    app.run()

if __name__ == '__main__':
    # Flaskサーバーを別のスレッドで起動
    server = threading.Thread(target=start_server)
    server.daemon = True
    server.start()
    time.sleep(1)  # サーバーが起動するのを待機

    # PyWebView GUIを起動
    webview.create_window('Excelファイル処理アプリ', 'http://127.0.0.1:5000', width=800, height=600)
    webview.start()